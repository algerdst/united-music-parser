import requests
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import glob
import os
import openpyxl
from datetime import datetime
import sys
from pytube import YouTube
import re


headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 YaBrowser/24.4.0.0 Safari/537.36'
}
# with open('output_path.txt', 'r', encoding='utf-8') as file:
#     for i in file:
#         save_path=i
# print(save_path)

def make_description(item_title):
    with open('description_text.txt', 'r', encoding='utf-8-sig') as file:
        for i in file:
            description_text=i.replace('item_title', item_title)
    return description_text

title_text=''
with open('title_text.txt', 'r', encoding='utf-8-sig') as file:
    for i in file:
        title_text = i

file = []
path = os.getcwd()
for filename in glob.glob(os.path.join(path, '*.xlsx')):
    file.append(filename)
filename = file[0]
book = openpyxl.load_workbook(filename)
sheet = book.active

target_url=input('Введите ссылку')

def get_links(url):
    links=[]
    with webdriver.Chrome() as browser:
        browser.get(url)
        time.sleep(5)
        max_page=int(browser.find_elements(By.CSS_SELECTOR, 'a.pager-page')[-2].text)
        print(f'Найдено страниц - {max_page}')

        for i in range(max_page):
            next_page = browser.find_elements(By.CSS_SELECTOR, 'a.pager-page')[-1]
            blocks = browser.find_element(By.ID,'list-products').find_elements(By.TAG_NAME, 'div')
            for block in blocks:
                try:
                    item_link=block.find_element(By.TAG_NAME, 'a').get_attribute('href')
                    if item_link not in links and item_link.endswith('.html'):
                        links.append(item_link)
                except:
                    pass
            next_page.click()
            time.sleep(5)
    return links

links=get_links(target_url)
print(f'Всего товаров по этому фильтру - {len(links)}')
count = 0
row = 2
with webdriver.Chrome() as browser:
    for url in links:
        response=requests.get(url, headers=headers).text
        soup=BeautifulSoup(response, 'lxml')
        title=soup.find('h1').text.strip()
        price=soup.find('span', class_='price').text.strip().split('р.')[0].replace(' ','')
        try:
            features_table=soup.find('table').find_all('tr')
            features=''
            for feature in features_table:
                td=feature.findAll('td')
                for i in td:
                    if len(td[1].text)>1:
                        features+=i.text+' - '
                    else:
                        features += i.text
                features=features[0:-2]
                features+='\n'
        except:
            features=''
        try:
            features_table = soup.find('div', class_='product__description').find('ul').find_all('li')
            for feature in features_table:
                features+=feature.text+'\n'
        except:
            pass
        article=soup.find('div', class_='product-details__description').find('b').text
        images_list=soup.find('ul', class_='product-bxslider').findAll('li')
        images=['https://united-music.ru/'+image.findNext('a')['href'] for image in images_list]
        images = ' | '.join(images)
        video_youtube_link=''
        yotube_request = 'https://www.youtube.com/results?search_query=' + title.replace(' ','+').replace('&', '%26')
        browser.get(yotube_request)
        video_youtube_links = browser.find_elements(By.ID, 'video-title')
        for youtube_link in video_youtube_links:
            youtube_link_title = youtube_link.text
            if title.lower() in youtube_link_title.lower():
                video_youtube_link = youtube_link.get_attribute('href')
                break
        description = make_description(title)
        features=description+'\n'+features
        title = title_text + ' '+title
        sheet.cell(column=3, row=row).value = article
        sheet.cell(column=4, row=row).value = video_youtube_link
        sheet.cell(column=5, row=row).value = price
        sheet.cell(column=13, row=row).value = title
        sheet.cell(column=15, row=row).value = description
        sheet.cell(column=20, row=row).value = images
        row += 1
        book.save(filename)
        count += 1
        print(f'Осталось собрать ссылок {len(links) - count}')


# скачивание видео с ютуба
# path = os.getcwd()
# file = []
# for filename in glob.glob(os.path.join(path, '*.xlsx')):
#     file.append(filename)
# filename = file[0]
# book = openpyxl.load_workbook(filename)
# sheet = book.active
# for i in range(2, sheet.max_row):
#     youtube_link=sheet[i][3].value
#     if youtube_link:
#         print(youtube_link)
#         try:
#             yt = YouTube(youtube_link)
#             stream = yt.streams.get_highest_resolution()
#             stream.download(save_path)
#             print('Видео загружено')
#         except Exception as e:
#             a = str(e)
#             print(a)
#             continue



